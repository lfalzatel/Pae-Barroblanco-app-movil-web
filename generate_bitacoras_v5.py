import openpyxl
from datetime import datetime, timedelta
import os

# --- DATA ---
STUDENT_NAME = "LUIS FERNANDO ALZATE LOPEZ"
STUDENT_ID = "75.107.425"
STUDENT_PHONE = "310 4604592"
STUDENT_EMAIL = "luis_falzate@sena.edu.co"

ENTERPRISE = "ALCALDÍA DE RIONEGRO"
BOSS_NAME = "GLORIA PATRICIA HINCAPIÉ YEPES (Secretaría Administrativa)"
BOSS_PHONE = "5204060"
BOSS_EXT = "2306"
BOSS_EMAIL = "serviciosadministrativos@rionegro.gov.co"

INSTRUCTOR_NAME = "CATHALINA CEBALLOS TORO"

START_DATE = datetime(2025, 12, 5)
END_DATE = datetime(2026, 5, 31)

TEMPLATE_PATH = r"c:\6. Viaje a san andres\3.Sistema-pae-barroblanco v-vercel-Converting to Progressive Web App (PWA)\1. Documentos practica\Formato Bitacora Actividades Prácticas.xlsx"
OUTPUT_DIR = r"c:\6. Viaje a san andres\3.Sistema-pae-barroblanco v-vercel-Converting to Progressive Web App (PWA)\1. Documentos practica\bitacoras_generadas"

CELLS = {
    'EMPRESA': 'B10',
    'BOSS_NAME': 'B13',
    'BOSS_PHONE': 'K13',
    'BOSS_EXT': 'N13',
    'BOSS_EMAIL': 'O13',
    'STUDENT_NAME': 'B16',
    'STUDENT_ID': 'J16',
    'STUDENT_PHONE': 'K16',
    'STUDENT_EMAIL': 'O16',
    'INSTRUCTOR': 'B46', # Overwriting Henry Mejia
    'NUM_BITACORA': 'C19',
    'PERIOD_START': 'D21',
    'PERIOD_END': 'H21',
}

# Activity rows: The template uses a 2-row merge (25-26, 27-28, etc.)
ACTIVITY_ROWS = [25, 27, 29, 31, 33, 35, 37, 39, 41, 43]

THEMES = [
    "Análisis de requerimientos técnicos y definición del alcance del Sistema PAE.",
    "Diseño de la arquitectura de la base de datos y esquemas de API REST.",
    "Implementación del sistema de autenticación y gestión de roles de usuario.",
    "Desarrollo del módulo de registro y seguimiento de beneficiarios PAE.",
    "Integración de servicios web y optimización de consultas a la base de datos.",
    "Implementación de funcionalidades de Progressive Web App (PWA) y modo offline.",
    "Optimización de rendimiento frontend y carga selectiva de recursos estáticos.",
    "Pruebas de usabilidad, corrección de errores de interfaz y mejora de UX.",
    "Refactorización de componentes principales y estandarización de estilos CSS.",
    "Desarrollo de paneles de control, reportes estadísticos y analíticas de datos.",
    "Pruebas de integración, despliegue en staging y configuración de Vercel.",
    "Documentación técnica final, manuales de usuario y cierre de fase de desarrollo."
]

def safe_write(ws, cell_coord, value):
    cell = ws[cell_coord]
    # Check if cell is part of a merged range
    for merged_range in ws.merged_cells.ranges:
        if cell_coord in merged_range:
            # Write to the top-left cell of the range
            master_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
            master_cell.value = value
            return
    cell.value = value

def generate():
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)

    curr_start = START_DATE
    num = 1

    while curr_start < END_DATE:
        curr_end = curr_start + timedelta(days=14)
        if curr_end > END_DATE:
            curr_end = END_DATE
        
        filename = os.path.join(OUTPUT_DIR, f"Bitacora_{num:02d}_{curr_start.strftime('%Y%m%d')}.xlsx")
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
        ws = wb.active
        
        # Header/Contact Info
        safe_write(ws, CELLS['EMPRESA'], ENTERPRISE)
        safe_write(ws, CELLS['BOSS_NAME'], BOSS_NAME)
        safe_write(ws, CELLS['BOSS_PHONE'], BOSS_PHONE)
        safe_write(ws, CELLS['BOSS_EXT'], BOSS_EXT)
        safe_write(ws, CELLS['BOSS_EMAIL'], BOSS_EMAIL)
        
        safe_write(ws, CELLS['STUDENT_NAME'], STUDENT_NAME)
        safe_write(ws, CELLS['STUDENT_ID'], STUDENT_ID)
        safe_write(ws, CELLS['STUDENT_PHONE'], STUDENT_PHONE)
        safe_write(ws, CELLS['STUDENT_EMAIL'], STUDENT_EMAIL)
        
        safe_write(ws, CELLS['INSTRUCTOR'], INSTRUCTOR_NAME)
        safe_write(ws, CELLS['NUM_BITACORA'], num)
        safe_write(ws, CELLS['PERIOD_START'], curr_start.strftime("%d/%m/%Y"))
        safe_write(ws, CELLS['PERIOD_END'], curr_end.strftime("%d/%m/%Y"))
        
        # Activities (Unique per bitácora)
        theme = THEMES[num-1] if (num-1) < len(THEMES) else "Continuación de actividades."
        sub_acts = [
            f"{theme} - Fase de planificación y análisis.",
            f"{theme} - Desarrollo y codificación técnica.",
            f"{theme} - Pruebas y validación de funcionalidades.",
            f"{theme} - Documentación y ajustes de calidad."
        ]
        
        days_total = (curr_end - curr_start).days
        step = max(1, days_total // len(sub_acts))
        
        for i, act in enumerate(sub_acts):
            if i >= len(ACTIVITY_ROWS): break
            row = ACTIVITY_ROWS[i]
            s_date = curr_start + timedelta(days=i*step)
            e_date = s_date + timedelta(days=step-1) if i < len(sub_acts)-1 else curr_end
            
            # Master cells for the 2-row merges:
            # Activity: B, Start Date: K, End Date: M
            safe_write(ws, f"B{row}", act)
            safe_write(ws, f"K{row}", s_date.strftime("%d/%m/%Y"))
            safe_write(ws, f"M{row}", e_date.strftime("%d/%m/%Y"))

        wb.save(filename)
        print(f"Generated: {filename}")
        
        curr_start = curr_end + timedelta(days=1)
        num += 1

if __name__ == "__main__":
    generate()
