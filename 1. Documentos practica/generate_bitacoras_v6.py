import openpyxl
from datetime import datetime, timedelta
import os

# --- DATOS OFICIALES ---
STUDENT_NAME = "LUIS FERNANDO ALZATE LOPEZ"
STUDENT_ID = "75.107.425"
STUDENT_PHONE = "310 4604592"
STUDENT_EMAIL = "luis_falzate@sena.edu.co"

ENTERPRISE = "ALCALDÍA DE RIONEGRO"
BOSS_NAME = "GLORIA PATRICIA HINCAPIÉ YEPES (Secretaría Administrativa)" # Jefa de Nelson
BOSS_PHONE = "5204060" # Ojo con esto si es otro, dejar el mismo
BOSS_EXT = "2306"
BOSS_EMAIL = "serviciosadministrativos@rionegro.gov.co"

INSTRUCTOR_NAME = "CATHALINA CEBALLOS TORO" # Nueva instructora SENA

START_DATE = datetime(2025, 12, 5) # CONFIRMADO POR EL USUARIO
END_DATE = datetime(2026, 5, 31)

TEMPLATE_PATH = r"c:\6. Mis proyectos PWA\3.Sistema-pae-barroblanco v-vercel-Converting to Progressive Web App (PWA)\1. Documentos practica\Formato Bitacora Actividades Prácticas.xlsx"
OUTPUT_DIR = r"c:\6. Mis proyectos PWA\3.Sistema-pae-barroblanco v-vercel-Converting to Progressive Web App (PWA)\1. Documentos practica\bitacoras_finales"

CELLS = {
    'EMPRESA': 'B10',
    'BOSS_NAME': 'B13',
    'BOSS_PHONE': 'K13',
    'BOSS_EXT': 'N13',
    'BOSS_EMAIL': 'O13',
    'STUDENT_NAME': 'B16',
    'STUDENT_ID': 'J16',
    'STUDENT_PHONE': 'K16',
    'STUDENT_EMAIL': 'O16',
    'INSTRUCTOR': 'B46', 
    'NUM_BITACORA': 'C19',
    'PERIOD_START': 'D21',
    'PERIOD_END': 'H21',
}

ACTIVITY_ROWS = [25, 27, 29, 31, 33, 35, 37, 39, 41, 43]

# --- ACTIVIDADES REALES PARA CADA BITÁCORA ---
# Cada llave (1-12) es el número de bitácora y tiene 4 o 5 actividades.
ACTIVITIES_PER_BITACORA = {
    1: [
        "Reunión de inducción y levantamiento de requerimientos iniciales del Sistema PAE.",
        "Definición de alcance, roles de usuario y flujos de trabajo del aplicativo.",
        "Configuración del entorno de desarrollo: Next.js, Tailwind CSS y Supabase.",
        "Diseño inicial de la arquitectura de la base de datos (tablas de perfiles y roles).",
        "Creación del repositorio en GitHub y configuración del control de versiones."
    ],
    2: [
        "Diseño de prototipos visuales y wireframes para la interfaz de usuario.",
        "Creación e implementación de los componentes base de la UI (botones, inputs).",
        "Configuración de Supabase Auth para el inicio de sesión básico con correo y contraseña.",
        "Implementación del inicio de sesión con Google (OAuth) en la pantalla principal.",
        "Pruebas de conectividad entre el frontend y la base de datos en Supabase."
    ],
    3: [
        "Desarrollo de las políticas de seguridad (RLS) en Supabase para proteger los datos.",
        "Creación del Layout principal del Dashboard con menú lateral responsivo.",
        "Implementación de la gestión del estado de sesión del usuario en React.",
        "Desarrollo del menú de navegación móvil (Bottom Navigation Bar).",
        "Ajustes de estilos y corrección de bugs visuales en la cabecera (Header)."
    ],
    4: [
        "Desarrollo del módulo de registro diario para la entrega de complementos PAE.",
        "Creación de la tabla de estudiantes y beneficiarios en la base de datos.",
        "Implementación de filtros de búsqueda por grado y grupo en la interfaz de registro.",
        "Desarrollo de la validación para evitar registros duplicados en el mismo día.",
        "Pruebas de inserción masiva de datos y optimización de consultas SQL."
    ],
    5: [
        "Desarrollo de la vista de 'Historial' para consultar entregas pasadas.",
        "Implementación de alertas y notificaciones visuales (Toasts) para el usuario.",
        "Optimización de consultas cruzadas entre usuarios y registros PAE.",
        "Creación del sistema multi-cuenta: almacenamiento local de sesiones paralelas.",
        "Desarrollo del selector de cuentas en el menú de perfil."
    ],
    6: [
        "Conversión de la aplicación web a Progressive Web App (PWA).",
        "Configuración del archivo manifest.json con íconos, colores y metadatos.",
        "Implementación del Service Worker para soporte básico offline y cacheo de rutas.",
        "Desarrollo de la interfaz (Prompt) para invitar a instalar la app en móviles.",
        "Pruebas de instalación de la PWA en dispositivos Android e iOS."
    ],
    7: [
        "Desarrollo del módulo de panel de control (Dashboard) para administradores.",
        "Implementación de tarjetas de resumen estadístico (raciones entregadas hoy, cupos).",
        "Creación de gráficas para visualizar el porcentaje de asistencia y entrega del PAE.",
        "Optimización de la carga de datos estadísticos usando funciones RPC en Supabase.",
        "Refactorización del código de las gráficas para mejorar el rendimiento visual."
    ],
    8: [
        "Integración del módulo de biometría (Huella dactilar y FaceID) mediante WebAuthn.",
        "Configuración de la verificación de credenciales biométricas en login.",
        "Ajustes en el formulario de login para hacer la sección de credenciales contraíble.",
        "Corrección de errores reportados en la navegación entre roles.",
        "Pruebas de usabilidad del login optimizado en dispositivos móviles."
    ],
    9: [
        "Desarrollo del módulo de gestión de usuarios para la secretaría y administradores.",
        "Implementación de la funcionalidad para asignar o cambiar roles.",
        "Mejora del sistema de advertencias para evitar sobrescritura de datos del PAE.",
        "Creación de ventanas modales (Popups) para confirmar acciones críticas.",
        "Pruebas de permisos RLS asegurando que los roles vean lo correspondiente."
    ],
    10: [
        "Desarrollo del submódulo de auditoría para rastrear acciones de usuarios.",
        "Optimización global de estilos usando Tailwind CSS para mejorar diseño.",
        "Ajuste de scroll y alturas máximas en los menús desplegables del perfil.",
        "Implementación del soporte para modo claro (Light mode) y modo oscuro (Dark mode).",
        "Revisión de accesibilidad y contrastes de color en la aplicación."
    ],
    11: [
        "Preparación del entorno de producción y configuración de variables de entorno.",
        "Despliegue de la aplicación en la plataforma Vercel.",
        "Pruebas de rendimiento de la aplicación desplegada.",
        "Simulación de carga con múltiples usuarios registrando datos simultáneamente.",
        "Corrección de errores detectados en el entorno de producción."
    ],
    12: [
        "Elaboración de la documentación técnica del proyecto y manuales de usuario.",
        "Creación de videotutoriales o guías rápidas para docentes y administrativos.",
        "Socialización de la herramienta con el líder de práctica y usuarios piloto.",
        "Ajustes finales de retroalimentación recibidos por la Alcaldía de Rionegro.",
        "Cierre de la fase de desarrollo, entrega de código fuente."
    ]
}

def safe_write(ws, cell_coord, value):
    cell = ws[cell_coord]
    for merged_range in ws.merged_cells.ranges:
        if cell_coord in merged_range:
            master_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
            master_cell.value = value
            return
    cell.value = value

def generate():
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)

    curr_start = START_DATE
    num = 1

    while curr_start < END_DATE and num <= 12:
        curr_end = curr_start + timedelta(days=14)
        if curr_end > END_DATE:
            curr_end = END_DATE
        
        filename = os.path.join(OUTPUT_DIR, f"Bitacora_Final_{num:02d}_{curr_start.strftime('%Y%m%d')}.xlsx")
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
        ws = wb.active
        
        # Header Info
        safe_write(ws, CELLS['EMPRESA'], ENTERPRISE)
        safe_write(ws, CELLS['BOSS_NAME'], BOSS_NAME)
        safe_write(ws, CELLS['BOSS_PHONE'], BOSS_PHONE)
        safe_write(ws, CELLS['BOSS_EXT'], BOSS_EXT)
        safe_write(ws, CELLS['BOSS_EMAIL'], BOSS_EMAIL)
        
        safe_write(ws, CELLS['STUDENT_NAME'], STUDENT_NAME)
        safe_write(ws, CELLS['STUDENT_ID'], STUDENT_ID)
        safe_write(ws, CELLS['STUDENT_PHONE'], STUDENT_PHONE)
        safe_write(ws, CELLS['STUDENT_EMAIL'], STUDENT_EMAIL)
        
        safe_write(ws, CELLS['INSTRUCTOR'], INSTRUCTOR_NAME)
        safe_write(ws, CELLS['NUM_BITACORA'], num)
        safe_write(ws, CELLS['PERIOD_START'], curr_start.strftime("%d/%m/%Y"))
        safe_write(ws, CELLS['PERIOD_END'], curr_end.strftime("%d/%m/%Y"))
        
        # Insert Activities
        activities = ACTIVITIES_PER_BITACORA.get(num, ["Continuación de actividades."])
        
        days_total = (curr_end - curr_start).days
        step = max(1, days_total // len(activities))
        
        for i, act in enumerate(activities):
            if i >= len(ACTIVITY_ROWS): break
            row = ACTIVITY_ROWS[i]
            s_date = curr_start + timedelta(days=i*step)
            e_date = s_date + timedelta(days=step-1) if i < len(activities)-1 else curr_end
            
            safe_write(ws, f"B{row}", act)
            safe_write(ws, f"K{row}", s_date.strftime("%d/%m/%Y"))
            safe_write(ws, f"M{row}", e_date.strftime("%d/%m/%Y"))

        wb.save(filename)
        print(f"Generada correctamente: {filename}")
        
        curr_start = curr_end + timedelta(days=1)
        num += 1

if __name__ == "__main__":
    generate()
