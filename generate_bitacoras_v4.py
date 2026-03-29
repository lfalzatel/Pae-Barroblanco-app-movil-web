import openpyxl
from datetime import datetime, timedelta
import os

# --- DATA ---
STUDENT_NAME = "LUIS FERNANDO ALZATE LOPEZ"
STUDENT_ID = "75.107.425"
ENTERPRISE = "ALCALDÍA DE RIONEGRO"
BOSS_NAME = "GLORIA PATRICIA HINCAPIÉ YEPES"
INSTRUCTOR_NAME = "CATHALINA CEBALLOS TORO"
START_DATE = datetime(2025, 12, 5)
END_DATE = datetime(2026, 5, 31)

TEMPLATE_PATH = r"c:\6. Viaje a san andres\3.Sistema-pae-barroblanco v-vercel-Converting to Progressive Web App (PWA)\1. Documentos practica\Formato Bitacora Actividades Prácticas.xlsx"
OUTPUT_DIR = r"c:\6. Viaje a san andres\3.Sistema-pae-barroblanco v-vercel-Converting to Progressive Web App (PWA)\bitacoras_generadas"
GIT_HISTORY_PATH = r"c:\tmp\git_history.txt"

CELLS = {
    'EMPRESA': 'B10',
    'RECTOR': 'B13',
    'NOMBRE_APRENDIZ': 'B16',
    'ID_APRENDIZ': 'J16',
    'INSTRUCTOR': 'B48',
    'NUM_BITACORA': 'C19',
    'FECHA_INICIO': 'D21',
    'FECHA_FIN': 'H21',
}

# Activity rows B25-B44 (ranges)
ACTIVITY_ROWS = range(25, 45)

# --- UTILS ---
def safe_write(ws, cell_coord, value):
    cell = ws[cell_coord]
    for merged_range in ws.merged_cells.ranges:
        if cell_coord in merged_range:
            master_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
            master_cell.value = value
            return
    cell.value = value

def get_git_activities():
    acts = {}
    if os.path.exists(GIT_HISTORY_PATH):
        with open(GIT_HISTORY_PATH, 'r', encoding='utf-8') as f:
            for line in f:
                parts = line.split(' ', 1)
                if len(parts) == 2:
                    date_str, msg = parts
                    acts.setdefault(date_str.strip(), []).append(msg.strip())
    return acts

def get_activities_for_period(start, end, git_acts, bitacora_num):
    if bitacora_num <= 2:
        return [
            "Análisis de requerimientos técnicos del proyecto Sistema PAE.",
            "Configuración del entorno de desarrollo y repositorio Git.",
            "Documentación de procesos de autenticación y seguridad.",
            "Desarrollo de componentes iniciales para gestión de estudiantes.",
            "Pruebas unitarias de flujo de registro."
        ]
    
    # Extract from git history
    period_acts = []
    curr = start
    while curr <= end:
        ds = curr.strftime("%Y-%m-%d")
        if ds in git_acts:
            period_acts.extend(git_acts[ds])
        curr += timedelta(days=1)
    
    if not period_acts:
        return ["Mantenimiento preventivo y optimización de código.", "Documentación técnica de módulos existentes."]
    
    # Unique and limit to 10 for display
    unique_acts = list(dict.fromkeys(period_acts))
    return unique_acts[:10]

# --- MAIN ---
if not os.path.exists(OUTPUT_DIR):
    os.makedirs(OUTPUT_DIR)

git_acts = get_git_activities()
periods = []
curr_start = START_DATE
num = 1

while curr_start < END_DATE:
    curr_end = curr_start + timedelta(days=14)
    if curr_end > END_DATE:
        curr_end = END_DATE
    
    filename = os.path.join(OUTPUT_DIR, f"Bitacora_{num:02d}_{curr_start.strftime('%Y%m%d')}.xlsx")
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb.active
    
    # Header
    safe_write(ws, CELLS['EMPRESA'], ENTERPRISE)
    safe_write(ws, CELLS['RECTOR'], BOSS_NAME)
    safe_write(ws, CELLS['NOMBRE_APRENDIZ'], STUDENT_NAME)
    safe_write(ws, CELLS['ID_APRENDIZ'], STUDENT_ID)
    safe_write(ws, CELLS['INSTRUCTOR'], INSTRUCTOR_NAME)
    safe_write(ws, CELLS['NUM_BITACORA'], num)
    safe_write(ws, CELLS['FECHA_INICIO'], curr_start.strftime("%d/%m/%Y"))
    safe_write(ws, CELLS['FECHA_FIN'], curr_end.strftime("%d/%m/%Y"))
    
    # Activities
    acts = get_activities_for_period(curr_start, curr_end, git_acts, num)
    for i, act in enumerate(acts):
        if i < len(ACTIVITY_ROWS):
            row = ACTIVITY_ROWS[i]
            safe_write(ws, f"B{row}", act)
    
    wb.save(filename)
    print(f"Generated: {filename}")
    
    curr_start = curr_end + timedelta(days=1)
    num += 1
